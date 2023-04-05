import datetime
import random
import openpyxl
import os
import inspect
import math

# Hello! this is an exercise log to show effort at a glance.
# For effect, have a scroll through faster than the eye can see  :)

def main():

    print(f"Hello! There are {task_count()} tasks here")

    paying_diner()
    tip_calculator()
    bmi_calculator()
    leapyear_checker()
    toss_a_coin()
    example_nested_lists()
    bury_the_treasure()
    rock_paper_scissors()
    student_avg_height()
    largest_num_without_max()
    fizzbuzz()
    password_generator()
    word_guessing()
    word_guessing_iteration()
    programming_coach()
    prime_checker()
    ceasar_cipher_experimentation()
    caesar_cipher_product()
    task_count()
    student_grading()
    dictionary_practice()
    blind_auction()
    days_in_month()
    the_infamous_calculator()
    calculator = Calculator()  # object-oriented-test
    calculator.run()  # object-oriented-test
    blackjack()
    number_guessing()
    debug_practice()
    higher_lower()
    coffee_shop()
    testing_image_processing()


def bury_the_treasure():
    row1 = ["", "️", ""]
    row2 = ["", "", "️"]
    row3 = ["", "", ""]
    map = [row1, row2, row3]
    print(f"{row1}\n{row2}\n{row3}")
    position = input("Where do you want to put the treasure? ")

    treasure_x = int(position[0]) - 1
    treasure_y = int(position[1]) - 1
    map[treasure_x][treasure_y] = "X"

    print(f"{row1}\n{row2}\n{row3}")


def tip_calculator():
    bill = float(input("how much was the bill?\n"))
    people = int(input("how many of you we're there?\n"))
    tip = float(input("what % tip would you like to give?\n").replace("%", ""))

    result = round(bill / people * (tip / 100), 2)
    print("{:.2f}".format(result))


def bmi_calculator():
    height = float(input("enter your height in m: "))
    weight = float(input("enter your weight in kg: "))

    result = round(weight / (height * height))

    if result < 18.5:
        print(f"Your BMI is {result}, you are underweight.")
    elif result < 25:
        print(f"Your BMI is {result}, you are normal weight.")
    elif result < 30:
        print(f"Your BMI is {result}, you are slightly overweight.")
    elif result < 35:
        print(f"Your BMI is {result}, you are obese.")
    else:
        print(f"Your BMI is {result}, you are clinically obese.")


def leapyear_checker():
    year = int(input("Which year do you want to check? "))

    if year % 4 == 0 and year % 100 != 0 or year % 4 == 0 and year % 100 == 0 and year % 400 == 0:
        print("Leap year")
    else:
        print("Not leap year")


def love_calculator():
    print("Welcome to the Love Calculator!")
    name1 = input("What is your name? \n")
    name2 = input("What is their name? \n")

    names = (name1 + name2)
    names = names.upper()

    score1 = 0
    score2 = 0

    score1 += names.count('T')
    score1 += names.count('R')
    score1 += names.count('U')
    score1 += names.count('E')

    score2 += names.count('L')
    score2 += names.count('O')
    score2 += names.count('V')
    score2 += names.count('E')

    result = str(score1) + str(score2)
    print(f"Your score is {result}.")


def toss_a_coin():
    toss = random.random()
    if toss < 0.5:
        print("Got tails bro!")
    else:
        print("Got heads broh!")


def paying_diner():
    names = input("gimme da names separated by commas only\n")
    each_name = names.split(",")
    count_names = len(each_name)
    pick_random = random.randint(0, count_names - 1)
    print(f"{each_name[pick_random]} pays broh!")


def example_nested_lists():
    fruits = ["Strawberries", "Nectarines", "Apples",
              "Grapes", "Peaches", "Cherries", "Pears"]
    vegetables = ["Spinach", "Kale", "Tomatoes", "Celery", "Potatoes"]

    dirty_dozen = [fruits, vegetables]

    print(dirty_dozen[1][1])


def rock_paper_scissors():

    rock = '''
        _______
    ---'   ____)
          (_____)
          (_____)
          (____)
    ---.__(___)
    '''

    paper = '''
        __________
    ---'      ____)
            ______)
            _______)
            _______)
    ---.__________)
    '''

    scissors = '''
        _______
    ---'    ____)____
               ______)
          __________)
         (____)
    ---.__(___)
    '''

    outcome = 0
    printables = ["ROCK", "PAPER", "SCISSORS"], [rock, paper, scissors]
    computer_choice = random.randint(0, 2)
    player_choice = int(
        input("enter 0 for rock, 1 for paper, 2 for scissors\n"))

    if computer_choice == 0 and player_choice == 2:
        outcome = "You lost!"
    elif computer_choice == 2 and player_choice == 0:
        outcome = "You won!"
    elif computer_choice < player_choice:
        outcome = "You won!"
    elif computer_choice == player_choice:
        outcome = "It's a draw!"
    elif computer_choice == 2 and player_choice == 1:
        outcome = "You lost!"
    else:
        print("throwException!")
        return

    print(f"The computer chose {printables[0][computer_choice]}!\n" +
          printables[1][computer_choice] + f"\nYou chose {printables[0][player_choice]}{printables[1][player_choice]}!\n" + outcome)


def student_avg_height():
    student_heights = input("Input a list of student heights ").split()
    for n in range(0, len(student_heights)):
        student_heights[n] = int(student_heights[n])

    count_of_heights = 0
    sum_of_heights = 0

    for aheight in student_heights:
        count_of_heights += 1
        sum_of_heights += int(aheight)

    result = round(sum_of_heights / count_of_heights)

    print(result)


def largest_num_without_max():
    max = 0
    nums = input(
        "Input a list of student scores separated by a comma").split(",")
    for num in nums:
        if int(num) > max:
            max = int(num)
    print(max)


def fizzbuzz():
    for num in range(1, 16):
        if num % 5 == 0 and num % 3 == 0:
            print("FizzBuzz!")
        elif num % 3 == 0:
            print("Fizz!")
        elif num % 5 == 0:
            print("Buzz!")
        else:
            print(num)


def password_generator():
    # import random
    letters = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y',
               'z', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    numbers = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']
    symbols = ['!', '#', '$', '%', '&', '(', ')', '*', '+']

    print("Welcome to the PyPassword Generator!")
    nr_letters = int(
        input("How many letters would you like in your password?\n"))
    nr_symbols = int(input(f"How many symbols would you like?\n"))
    nr_numbers = int(input(f"How many numbers would you like?\n"))

    password = ""

    while nr_letters + nr_symbols + nr_numbers != 0:
        randomize = random.randint(0, 2)
        if randomize == 0 and nr_letters > 0:
            pass_letter = letters[random.randint(0, len(letters)-1)]
            password += pass_letter
            nr_letters -= 1
        elif randomize == 1 and nr_symbols > 0:
            pass_symbol = symbols[random.randint(0, len(symbols)-1)]
            password += pass_symbol
            nr_symbols -= 1
        elif randomize == 2 and nr_numbers > 0:
            pass_numbers = numbers[random.randint(0, len(numbers)-1)]
            password += str(pass_numbers)
            nr_numbers -= 1
        else:
            continue

    print(password)

# region reborg
# def pass_a_hurdle():
    # turn_left()
    # move()
    # turn_left()
    # turn_left()
    # turn_left()
    # move()
    # turn_left()
    # turn_left()
    # turn_left()
    # move()
    # turn_left()

# while not at_goal():
#     while front_is_clear():
#         move()
#     pass_a_hurdle()

# while not at_goal():
#     while front_is_clear():
#         move()
#         turn_left()
#         if front_is_clear():
#             move()
#             continue
#         turn_right()
#     if right_is_clear():
#         turn_right()
#         move()
#     if not front_is_clear() and not right_is_clear():
#         turn_left()

# while front_is_clear():
#     move()
# turn_left()
# while not at_goal():
#     if right_is_clear():
#         turn_right()
#         move()
#     elif front_is_clear():
#         move()
#     else:
#         turn_left()

# endregion

# Step 1


def programming_coach():
    # Open the spreadsheet (create it if it doesn't exist)
    try:
        workbook = openpyxl.load_workbook('answers.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Select the active sheet
    sheet = workbook.active

    # Ask the questions and save the answers
    last_row = sheet.max_row

    if last_row > 1:
        sheet.append([])

    question1 = input("What is your primary programming goal for today? ")
    sheet.append(["What is your primary programming goal for today?", question1,
                  datetime.date.today().strftime('%Y-%m-%d')])

    question2 = input(
        "What specific task are you working on right now, and how does it contribute to your overall goal? ")
    sheet.append(["What specific task are you working on right now, and how does it contribute to your overall goal?", question2,
                  datetime.date.today().strftime('%Y-%m-%d')])

    question3 = input(
        "What challenges are you facing, and what resources or strategies can you use to overcome them? ")
    sheet.append(["What challenges are you facing, and what resources or strategies can you use to overcome them?", question3,
                  datetime.date.today().strftime('%Y-%m-%d')])

    question4 = input(
        "What did you learn today, and how can you apply it to future projects or tasks? ")
    sheet.append(["What did you learn today, and how can you apply it to future projects or tasks?", question4,
                  datetime.date.today().strftime('%Y-%m-%d')])

    question5 = input(
        "Did you make any mistakes today, and what can you learn from them? ")
    sheet.append(["Did you make any mistakes today, and what can you learn from them?", question5,
                  datetime.date.today().strftime('%Y-%m-%d')])
    # Save the spreadsheet
    workbook.save('answers.xlsx')

    os.startfile('answers.xlsx')


def word_guessing():

    word_list = ["aardvark", "baboon", "camel"]

    chosen_word = word_list[random.randint(0, 2)]

    guess = input("guess a letter!\n").lower()

    for letter in chosen_word:
        if letter == guess:
            print("a match!")
        else:
            print("dun dun duuuun!")


def word_guessing_iteration():
    from hangman_words import stages

    end_of_game = False
    lives = 6
    from hangman_words import word_list
    chosen_word = random.choice(word_list)

    print(f'Pssst, the solution is {chosen_word}.')

    display = []
    for letter in chosen_word:
        display += "_"
    print(display)

    while not end_of_game:
        guess = input("guess a letter: ").lower()
        os.system('cls')
        for index, letter in enumerate(chosen_word):
            if letter == guess:
                if display[index] == guess:
                    print("you already guessed this one!")
                    continue
                display[index] = guess
                print("got one!")

        if guess not in chosen_word:
            lives -= 1
            if lives == 0:
                print("You got smekt")
                end_of_game = True
        print(stages[lives])
        print(display)
        if "_" not in display:
            print("you win!")
            end_of_game = True


def prime_checker():
    number = int(input("Check this number: "))
    for divisor in range(2, number+1):
        if number != divisor and number % divisor == 0:
            print("not a prime number")
            return
    print("it's a prime number!")


def ceasar_cipher_experimentation():

    alphabet = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l',
                'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']
    direction = input(
        "Type 'encrypt' to encrypt, type 'decrypt' to decrypt:\n")
    text = input("Type your message:\n").lower()
    shift = int(input("Type the shift number:\n"))

    def encrypt(plain_text, shift_amount):
        encrypted = ""
        for letter in plain_text:
            if letter == " ":
                encrypted += " "
                continue
            position = alphabet.index(letter)
            if position + shift_amount > 25:
                new_position = (shift_amount - (26 - position))
            else:
                new_position = position + shift_amount
            encrypted += alphabet[new_position]
            mem = open('txt.txt', 'w')
            mem.write(encrypted)
            mem.close()
        print(f"The encrypted text is: {encrypted.replace(' ','')}")

    def decrypt():

        decrypted = ""
        letter_to_add = ""

        for letter in text:
            if letter == " ":
                decrypted += " "
                continue
            for alphabet_index, alphabet_letter in enumerate(alphabet):
                if letter == alphabet_letter:
                    letter_to_add = alphabet[alphabet_index - shift]
                    decrypted += letter_to_add
        print(f"The decrypted message is: {decrypted}")

    def decrypt_from_memory():

        mem = open('txt.txt', 'r')
        input_to_decrypt = mem.read()
        decrypted = ""
        letter_to_add = ""

        for letter in input_to_decrypt:
            if letter == " ":
                decrypted += " "
                continue
            for alphabet_index, alphabet_letter in enumerate(alphabet):
                if letter == alphabet_letter:
                    letter_to_add = alphabet[alphabet_index - shift]
                    decrypted += letter_to_add
        print(f"The decrypted message is: {decrypted}")

    if direction == "encrypt":
        encrypt(plain_text=text, shift_amount=shift)
    elif direction == "decrypt":
        decrypt()
    elif direction == "decrypt_from_memory":
        decrypt_from_memory()


def caesar_cipher_product():

    from art import logo
    main_input = "Y"
    print(logo)
    while main_input == "Y":
        alphabet = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l',
                    'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']

        def caesar(start_text, shift_amount, cipher_direction):
            end_text = ""
            if shift_amount > 26:
                shift_amount = shift_amount % 26
            if cipher_direction == "decode":
                shift_amount *= -1
            for char in start_text:
                if char not in alphabet:
                    end_text += char
                    continue
                position = alphabet.index(char)
                if position + shift_amount > 25:
                    new_position = shift_amount - (26 - position)
                else:
                    new_position = position + shift_amount
                end_text += alphabet[new_position]

            print(f"Here's the {cipher_direction}d result: {end_text}")

        direction = input(
            "Type 'encode' to encrypt, type 'decode' to decrypt:\n")
        text = input("Type your message:\n").lower()
        shift = int(input("Type the shift number:\n"))

        caesar(start_text=text, shift_amount=shift, cipher_direction=direction)

        redo = input("Go again? Y | N\n")
        if redo == "Y":
            continue
        else:
            main_input = "N"


def task_count():
    main_source = inspect.getsource(main)
    num_lines = len(main_source.split("\n"))
    return num_lines - 6


def student_grading():

    student_scores = {
        "Harry": 81,
        "Ron": 78,
        "Hermione": 99,
        "Draco": 74,
        "Neville": 62,
    }

    student_grades = {}

    for stud in student_scores:
        if student_scores[stud] > 90:
            student_grades = {stud: "Outstanding!"}
        elif student_scores[stud] > 80:
            student_grades = {stud: "Exceeds Expectations"}
        elif student_scores[stud] > 70:
            student_grades = {stud: "Acceptable"}
        else:
            student_grades = {stud: "Fail"}
        print(student_grades)


def dictionary_practice():

    # travel_log = {
    #     "Skellige": {"areas_visited": {"sauna": 6, "forgotten coast": 9, "garden's of death": 6}}
    # }
    # print(travel_log)

    travel_log = [
        {
            "faction": "Nilfgaard",
            "visits": 12,
            "cities": ["Nilfgaard", "Vizima", "Glyswen"]
        },
        {
            "faction": "Temeria",
            "visits": 5,
            "cities": ["Pontaria", "Sodden", "Maribor"]
        },
    ]

    def add_new_faction(travel_log, faction, visits, cities):

        travel_log += [{"faction": faction,
                       "visits": visits,
                        "cities": cities}]

    add_new_faction(travel_log, "Scoia'tael", 2, ["Mahakam", "Dol Blathanna"])
    print(travel_log)


def blind_auction():

    from art import gavel_logo
    print(gavel_logo)
    more_bidders = False
    start = input("Are you ready to start the blind auction?\n")
    if start == "YES":
        more_bidders = True
    bids = {}
    while more_bidders:
        os.system('cls')
        print(gavel_logo)
        name = input("enter your name\n")
        bid = input("enter your bid\n")
        bids[name] = bid
        next_bid = input("are there any more bidders? Y|N\n")
        if next_bid == "N":
            more_bidders = False
    winner = max(bids)
    print(f"The winner is {winner} with a bid of {bids[winner]}!")


def days_in_month():
    """Asks for a year and a month, and tells you how many days are in the month.
    If the year is a leap year, there is one extra day in february."""
    def is_leap(year):
        if year % 4 == 0:
            if year % 100 == 0:
                if year % 400 == 0:
                    return True
                else:
                    return False
            else:
                return True
        else:
            return False

    month_days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

    year = int(input("Enter a year: "))
    month = int(input("Enter a month: "))
    if (is_leap(year) and month == 1):
        result = month_days[month] + 1
    else:
        result = month_days[month]

    return result


def the_infamous_calculator():

    from art import calculator_logo
    print(calculator_logo)

    def add(n1, n2):
        return n1 + n2

    def substract(n1, n2):
        return n1 - n2

    def multiply(n1, n2):
        return n1 * n2

    def divide(n1, n2):
        return n1 / n2

    operations = {
        "+": add,
        "-": substract,
        "*": multiply,
        "/": divide,
        "e": exit,
    }

    answer = ''
    n1 = ''
    lets_calculate = True
    while lets_calculate == True:
        for operation in operations:
            print(operation)
        if n1 != '':
            n1 = answer
        else:
            n1 = float(input("What's the first number?: "))
        operation_symbol = input("Pick an operation")
        n2 = float(input("Whats the operating number?: "))

        answer = operations[operation_symbol](n1, n2)
        print(f"{n1} {operation_symbol} {n2} = {answer}")
        operation = input(
            "Enter to continue the same calculation. e to exit or n for new: ")
        if operation == 'n':
            the_infamous_calculator()
        if operation == 'e':
            lets_calculate = False

# testing an object-oriented version


class Calculator:
    def __init__(self):
        self.answer = ''
        self.operation_symbol = ''
        self.operations = {
            "+": self.add,
            "-": self.substract,
            "*": self.multiply,
            "/": self.divide,
            "e": exit,
        }

    @staticmethod
    def add(n1, n2):
        return n1 + n2

    @staticmethod
    def substract(n1, n2):
        return n1 - n2

    @staticmethod
    def multiply(n1, n2):
        return n1 * n2

    @staticmethod
    def divide(n1, n2):
        return n1 / n2

    def run(self):
        from art import calculator_logo
        print(calculator_logo)

        lets_calculate = True
        while lets_calculate == True:
            for operation in self.operations:
                print(operation)
            if not self.answer:
                n1 = float(input("What's the first number?: "))
            else:
                n1 = self.answer
            self.operation_symbol = input(
                "Pick an operation")
            if self.operation_symbol == 'e':
                self.operations[self.operation_symbol]()
            n2 = float(input("Whats the operating number?: "))
            self.answer = self.operations[self.operation_symbol](n1, n2)
            print(f"{n1} {self.operation_symbol} {n2} = {self.answer}")
            operation = input(
                "Press enter to continue the same calculation. e to exit or n for new: ")
            if operation == 'n':
                the_infamous_calculator()
            if operation == 'e':
                lets_calculate = False

    # calculator = Calculator()
    # calculator.run()


def blackjack():
    from art import blackjack_logo
    print(blackjack_logo)

    def draw_cards_to(hand):
        cards = [11, 2, 3, 4, 5, 6, 7, 8, 9, 10, 10, 10, 10]
        hand.append(cards[random.randint(0, 12)])

    def flip_ace_in(hand):
        ace = hand.index(11)
        hand[ace] = 1

    def draw_the_table():
        # print(sum(dealers_hand)
        # print(dealers_hand)
        print(table)
        print(your_hand)
        print(sum(your_hand))
        print("------------")

    def choose_actions():

        while sum(your_hand) < 21 or 11 in your_hand and sum(your_hand) != 21:

            draw_the_table()

            if 11 in your_hand:
                choice = input(
                    "to flip an ace send 'f' to hit send 'h', or just press enter to continue: ")
                if choice == 'f':
                    flip_ace_in(your_hand)
                    continue
                if choice == 'h':
                    draw_cards_to(your_hand)
                    continue
                else:
                    return
            else:
                choice = input(
                    "to hit send 'h' or just press enter to continue: ")
                if choice == 'h':
                    draw_cards_to(your_hand)
                    continue
            return
        else:
            return

    def end_scenario():
        draw_the_table()
        if sum(your_hand) == 21:
            print("A Blackjack!!")
        while sum(dealers_hand) < 17:
            draw_cards_to(dealers_hand)
            if sum(dealers_hand) > 21 and 11 in dealers_hand:
                flip_ace_in(dealers_hand)
            for i in range(len(dealers_hand) - len(table)):
                table.append("[ ]")
            if sum(dealers_hand) < 17:
                table.append("[ ]")
            for i in range(1, len(dealers_hand)):
                table[i] = (f"[{dealers_hand[i]}]")
            print("********************")
            input("ready to see the dealer's next card?")
            print("********************")
            print(sum(dealers_hand))
            draw_the_table()

    def calculate_score():
        print(f"the dealer got {sum(dealers_hand)}")
        print(f"you got {sum(your_hand)}")
        if sum(your_hand) > 21 and sum(dealers_hand) > 21:
            print("You both bust!")
        elif sum(dealers_hand) > 21 or sum(your_hand) <= 21 and sum(your_hand) > sum(dealers_hand):
            print("You won!")
        elif sum(your_hand) == sum(dealers_hand):
            print("It's a draw!")
        else:
            print("You lost!")

    dealers_hand = []
    draw_cards_to(dealers_hand)

    table = [f"[{dealers_hand[0]}]", "[ ]"]

    your_hand = []
    while len(your_hand) < 2:
        draw_cards_to(your_hand)

    choose_actions()
    end_scenario()
    calculate_score()


def number_guessing():

    from art import number_guessing_logo
    print(number_guessing_logo)
    secret_num = random.randint(1, 100)
    difficulty = [20, 15, 10, 5]
    print("Welcome to GUESS 1-100!")
    print("\nChoose the difficulty!")
    print("send 0 for easy: 20 guesses")
    print("send 1 for medium: 15 guesses")
    print("send 2 for hard: 10 guesses")
    print("send 3 for very hard: 5 guesses, yikes!")
    choice = int(input("\nsend 0,1,2 or 3!\n"))
    lives = difficulty[choice]
    for i in range(0, lives):
        print(f"lives left: {lives - i}")
        guess = int(input("out of 1-100, guess which number i've picked!\n"))
        if secret_num == guess:
            print("that's right, good job!!")
            break
        elif secret_num > guess:
            print("nope, higher!")
        elif secret_num < guess:
            print("nope, lower!")
    if guess != secret_num:
        print("bummer!")
    if input("Play again? y|n") == 'y':
        number_guessing()


def debug_practice():

    # the problem is that the range function starts and stops at -1 of the given values. so 0 and 19 here
    def my_function():
        for i in range(1, 21):  # fixed
            if i == 20:
                print("You got it")
    my_function()

    # Reproduce the Bug, the randint range was set at 1-6, but it works dissimilarly to the range function
    from random import randint
    dice_imgs = ["❶", "❷", "❸", "❹", "❺", "❻"]
    dice_num = randint(0, 5)
    print(dice_imgs[dice_num])

    # the code was missing a = sign, and ignored the year 1994
    year = int(input("What's your year of birth?"))
    if year > 1980 and year < 1994:
        print("You are a millenial.")
    elif year >= 1994:
        print("You are a Gen Z.")

    # the code was missing an (f"string) int("string") and an indentation
    age = int(input("How old are you?"))
    if age > 18:
        print(f"You can drive at age {age}.")

    # the == is a comparison operator, it needed to be switched into =
    pages = 0
    word_per_page = 0
    pages = int(input("Number of pages: "))
    word_per_page = int(input("Number of words per page: "))
    total_words = pages * word_per_page
    print(total_words)

    # there was an issue with the code's indentation, making the list receive only a single append
    def mutate(a_list):
        b_list = []
        for item in a_list:
            new_item = item * 2
            b_list.append(new_item)
        print(b_list)

    mutate([1, 2, 3, 5, 8, 13])


def higher_lower():

    from higher_lower_game_data import data
    from art import higher_lower_vs
    from art import higher_lower_logo

    def print_stats(ig_star, stats=None, stat_removal=None):
        """input ig_star_data, prints ['name', 'follower_count', 'description', 'country']  
        Optional param1: stats 
            specify which stats to include with a string or a list of strings, the rest won't be printed. 
        Optional param2: stat_removal 
            specify which stat to remove with a string or a list of strings, and print the rest.
            e.g 'no_follower_count'"""

        check = {
            'no_follower_count': 'follower_count',
            'no_name': 'name',
            'no_description': 'description',
            'no_country': 'country'}

        if stats == None:
            stats = list(check.values())

        if stat_removal != None:
            for item in check:
                if item in stat_removal:
                    stats.remove(check[item])
        for item in stats:
            print(ig_star[item], " ", end='')

        print()

    def game_screen():
        print_stats(ig_star_one, stat_removal='no_follower_count')
        print(higher_lower_vs)
        print()
        print_stats(ig_star_two, stat_removal='no_follower_count')
        print("-----------")

    print(higher_lower_logo)
    play = ''
    score = [0]
    while play == '':
        ig_star_one = data[random.randint(0, len(data)-1)]
        ig_star_two = data[random.randint(0, len(data)-1)]
        while ig_star_two == ig_star_one:
            ig_star_two = data[random.randint(1, len(data) - 1)]

        game_screen()

        choice = input(
            "\nWhich one has the higher follower count? send 1 or 2: ")

        os.system('cls')
        if os.name == 'posix':
            os.system('clear')

        print("*--------------------------------------*")
        print_stats(ig_star_one, ['name', 'follower_count'])
        print_stats(ig_star_two, ['name', 'follower_count'])

        win_condition_one = choice == '1' and ig_star_one[
            'follower_count'] > ig_star_two['follower_count']
        win_condition_two = choice == '2' and ig_star_one[
            'follower_count'] < ig_star_two['follower_count']

        if win_condition_one or win_condition_two:

            if math.fabs(float(ig_star_one['follower_count'] - ig_star_two['follower_count'])) < 10:
                score.append(50)
            else:
                score.append(10)
            print("*--------------------------------------*")
            print(f"you got it! {score[-1]} points!")
            print(f"your current score is {sum(score)}")
            print("*--------------------------------------*")
            play = input(
                "ready for the next pair? hit enter to continue, or send n to quit\n\n\n")
            os.system('cls')
        else:
            if math.fabs(float(ig_star_one['follower_count'] - ig_star_two['follower_count'])) > 20:
                pop_count = 2
            else:
                pop_count = 1
            if sum(score) > 0:
                print("*--------------------------------------*")
                print(f"\nbummer! you had {sum(score)} points.")

                if len(score) - pop_count > 0:
                    for i in range(0, pop_count):
                        print(f"**pop**you lost {score.pop()} points")
                else:
                    score.clear()

                print(f"now you have {sum(score)}")
                print("*--------------------------------------*")
            else:
                print("bummer!")
            play = input(
                "\nhit enter to continue, or send n to quit\n\n\n")
            os.system('cls')
            if os.name == 'posix':
                os.system('clear')


def coffee_shop():

    from coffee_data import MENU, resources, coin_types, units, cost

    def calculate_coins(coin_list):
        coin_list = list(map(int, coin_list))
        coin_value = [0.25, 0.10, 0.05, 0.01]
        for i in range(0, len(coin_list)):
            coin_list[i] = coin_list[i] * coin_value[i]
        return (sum(coin_list))

    def manage_resources(ingredient):

        for item in ingredient:
            if ingredient[item] > resources[item]:
                print("* internal screech *")
                return False
        return True

    def successfull_transaction(ingredient, choice):
        for item in ingredient:
            resources[item] -= ingredient[item]
        resources['money'] += cost[choice]
        return

    def get_report():
        os.system('cls')
        for item in resources:
            print(f"{item}: {resources[item]} {units[item]} ")
        input("done viewing?")
        ai_service()

    def order_more():
        more = input("more coffee? y|n\n")
        if more == 'report':
            get_report()
        if more == 'y':
            ai_service()
        else:
            exit()

    def ai_service():

        total_coins = []
        transaction_completed = False

        os.system('cls')
        print("**Welcome to the coffee shop!**")
        print()
        for index, item in enumerate(MENU):
            print(f"{index} {item}, {cost[item]} $")
        print()

        request = input("what would you like? (0, 1, 2): ")
        if request == 'report':
            get_report()
        else:
            choice = list(MENU)[int(request)]
            ingredient = MENU[choice]

        os.system('cls')
        if manage_resources(ingredient):
            print(f"{cost[choice]} $ for {choice} please\n")
        else:
            print("Sorry we've ran out of ingredients!\n")
            order_more()

        for coin in coin_types:
            coin = input(
                f"insert {coin}, or just hit enter\n")
            if coin == 'report':
                answer = input("why do you want a report now? just get your coffee!")
                if answer == 'report':
                    get_report()
                else:
                    ai_service()
            if coin != '':
                total_coins.append(coin)
                total_value = calculate_coins(total_coins)
                os.system('cls')
                print(f"you've inserted a total of {round(total_value, 2)} $")
                if total_value == cost[choice]:
                    print(f"that's exactly what's needed for {choice}")
                elif total_value > cost[choice]:
                    overhead = total_value - cost[choice]
                    print(f"that's {overhead} $ over, here's your change")
                if total_value == cost[choice] or total_value > cost[choice]:
                    print("Enjoy your coffee!")
                    transaction_completed = True
                    successfull_transaction(ingredient, choice)
                    break
                else:
                    print(f"that's {round(total_value - cost[choice], 2)} $ short, do you have some more coins?")

        if transaction_completed == False and total_coins:
            print(f"bummer, that's not enough, here are your coins: {round((total_value), 2)}")
        elif transaction_completed == False:
            print("\nyou have got to insert some coins, buddy")

        order_more()

    ai_service()

def testing_image_processing():
    import colorgram
    from PIL import Image, ImageFilter

    img = Image.open("turtleart2.png")
    img = img.filter(ImageFilter.GaussianBlur(radius=1))
    img.save("image_antialiased2.jpg")

    colors = colorgram.extract('fengzhu.jpg', 30)
    color_list = []
    for i in range(0, len(colors)):
        lets_color = colors[i]
        color_set = [lets_color.rgb.r, lets_color.rgb.maxsize, lets_color.rgb.g]
        color_list.append(color_set)
    print(color_list)

    for _ in range(1,100):
        random_color = []
        for i in range(3):
            random_color.append(random.randint(0,255))
        # draw.pencolor(random_color)
        # draw.circle(100)
        # draw.right(10)

def testing_drawing_libraries():

    FULL_CIRCLE = 360
    square = 4

    def print_row():
        for i in range(10):
            color_list = [
                [28, 33, 41],
                [74, 78, 100],
                [21, 32, 29],
                [25, 19, 27],
                [73, 118, 104],
                [118, 187, 164],
                [130, 136, 170],
                [87, 63, 100],
                [50, 53, 74],
                [104, 97, 151],
                [63, 38, 76],
                [174, 238, 223],
                [195, 221, 230],
                [18, 17, 15],
                [121, 72, 156],
                [98, 149, 139],
                [146, 106, 175],
                [158, 162, 215],
                [47, 75, 71],
                [151, 223, 206],
                [93, 164, 132],
                [181, 125, 216],
                [231, 186, 221],
                [46, 84, 65],
                [154, 231, 197],
                [84, 82, 80]
            ]

    # def turn_left():
    #     draw.dot(random.randint(minsize, maxsize))
    #     draw.left(90)
    #     draw.forward(distance)
    #     draw.left(90)


    # def turn_right():
    #     draw.dot(random.randint(minsize, maxsize))
    #     draw.right(90)
    #     draw.forward(distance)
    #     draw.right(90)

    multiplier = 0.24
    base = [150, 470]
    minsize = int(multiplier * base[0])
    maxsize = int(multiplier * base[1])
    distance = 120


    # for i in range(10):
    #     count_of_sides = square + i
    #     for i in range(count_of_sides):
            # draw.pencolor(random.randint(1,256),random.randint(1,256),random.randint(1,256))
            # draw.forward(50)
            # draw.right(FULL_CIRCLE / count_of_sides)
            # draw.forward(50)

    # for _ in range(200):
    #     random_color = []
    #     for i in range(3):
    #         random_color.append(random.randint(0,255))
    #     draw.pencolor(random_color)
    #     draw.forward(30)
    #     draw.setheading(random.choice(directions))
    #     draw.forward(30)
    # ---------------------------------------

    # screen = Screen()
    # screen.colormode()
    # screen.colormode(255)
    # screen.tracer(8,25)
    # screen.setup(2560, 1440)
    # draw = t()
    # draw.color("black")
    # draw.hideturtle()
    # draw.penup()
    # draw.pensize(20)
    # draw.speed(0)
    # draw.pencolor(random.choice(color_list))
    # draw.dot(random.randint(minsize, maxsize))
    # draw.forward(distance)

    # def set_starting_position():
    #     draw.right(90)
    #     draw.forward(500)
    #     draw.left(270)
    #     draw.forward(500)
    #     draw.left(180)


    # set_starting_position()
    # for i in range(5):
    #     print_row()
    #     turn_left()
    #     print_row()
    #     turn_right()

    # screen.exitonclick()
    # directions = [0, 90, 180, 270]

if __name__ == "__main__":
    main()
